import json
import os
import sys
import time
import copy
import socket
import platform
import uuid
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from dataclasses import dataclass, asdict
from typing import Optional, List
from openpyxl import load_workbook, Workbook
import csv

APP_TITLE = "Tönchtory"
SESSION_FILE = "tonch_inventory_fast_session.json"
AUDIT_FILE = "tonchtory_audit_enterprise.jsonl"
AUDIT_SUMMARY_FILE = "tonchtory_audit_summary.json"

HEADER_SYNONYMS = {
    "codigo": ["cod produto", "codigo produto", "código produto", "codigo", "código", "cod", "sku", "id produto", "product code"],
    "nome": ["nome do produto", "produto", "nome", "descricao", "descrição", "description", "item"],
    "marca": ["marca", "brand", "fabricante"],
    "quantidade": ["quantidade", "qtd", "qtde", "estoque", "saldo", "stock"],
    "disponivel": ["disponível", "disponivel", "saldo disponível", "saldo disponivel", "available"],
    "unidade": ["unidade", "und", "un", "unit", "u.m.", "um"],
    "contagem": ["contagem inventário", "contagem inventario", "contagem"],
    "diferenca": ["diferença", "diferenca"],
}

COLORS = {
    "bg": "#10141b",
    "panel": "#182130",
    "header": "#0f3d91",
    "border": "#2f3b4f",
    "text": "#f4f7fb",
    "muted": "#a8b5c7",
    "blue": "#2d89ef",
    "blue_dark": "#5aa7ff",
    "red": "#c62828",
    "white": "#202a3a",
    "alt": "#243044",
    "select_bg": "#2d89ef",
    "select_fg": "#ffffff",
    "tile": "#151d2b",
    "tile2": "#0f6fd6",
}

def tile_button(master, text, command, primary=False, danger=False):
    bg = COLORS["red"] if danger else (COLORS["blue"] if primary else COLORS["tile"])
    active = "#e53935" if danger else ("#4aa3ff" if primary else "#243044")
    return tk.Button(
        master,
        text=text,
        command=command,
        bg=bg,
        fg=COLORS["text"],
        activebackground=active,
        activeforeground=COLORS["select_fg"],
        relief="flat",
        bd=0,
        padx=14,
        pady=10,
        font=("Segoe UI", 9, "bold"),
        cursor="hand2"
    )

def center_window(win, width, height):
    win.update_idletasks()
    x = int((win.winfo_screenwidth() - width) / 2)
    y = int((win.winfo_screenheight() - height) / 2)
    win.geometry(f"{width}x{height}+{x}+{y}")

def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    mapping = str.maketrans("ãáàâéêíóõôúç", "aaaaeeiooouc")
    text = text.translate(mapping)
    return " ".join(text.split())

def normalize_code(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text

def to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0

def fmt_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".replace(".", ",")

@dataclass
class InventoryItem:
    codigo: str
    nome: str = ""
    marca: str = ""
    quantidade: float = 0.0
    disponivel: float = 0.0
    unidade: str = ""
    contagem_inventario: float = 0.0
    linha_excel: Optional[int] = None
    novo_item: bool = False
    contagem_sem_planilha: bool = False
    ultima_leitura: str = ""
    total_scans: int = 0

    @property
    def diferenca(self) -> float:
        return self.contagem_inventario - self.disponivel

    @property
    def status(self) -> str:
        if self.contagem_sem_planilha:
            return "NOVO"
        if self.contagem_inventario == 0:
            return "ZERADO"
        if self.diferenca == 0:
            return "OK"
        if self.diferenca > 0:
            return "SOBRA"
        return "FALTA"

    @property
    def origem(self) -> str:
        if self.contagem_sem_planilha:
            return "NOVA CONTAGEM"
        return "NOVO" if self.novo_item else "BASE"

    def sincronizar_contagem_sem_planilha(self):
        """
        Estoque criado dentro do sistema:
        - Nova contagem sem planilha: cada leitura atualiza Contagem, Quantidade e Disponível.
        - Planilha importada + item novo cadastrado: o item também nasce no sistema, então
          cada leitura atualiza Contagem, Quantidade e Disponível.
        - Itens originais da planilha importada seguem o fluxo normal.
        """
        if self.contagem_sem_planilha or self.novo_item:
            valor = max(0.0, float(self.contagem_inventario or 0.0))
            self.contagem_inventario = valor
            self.quantidade = valor
            self.disponivel = valor


class AuditManager:
    def __init__(self, base_dir: str):
        self.base_dir = base_dir
        self.path = os.path.join(base_dir, AUDIT_FILE)
        self.summary_path = os.path.join(base_dir, AUDIT_SUMMARY_FILE)
        self.session_id = uuid.uuid4().hex[:12]
        self.started_at = time.time()
        self.host = socket.gethostname()
        self.user = os.environ.get("USERNAME") or os.environ.get("USER") or ""
        self.ip = self._detect_ip()
        self.open_count = self._increment_open_count()

    def _detect_ip(self):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.connect(("8.8.8.8", 80))
            ip = sock.getsockname()[0]
            sock.close()
            return ip
        except Exception:
            try:
                return socket.gethostbyname(socket.gethostname())
            except Exception:
                return "indisponivel"

    def _load_summary(self):
        if not os.path.exists(self.summary_path):
            return {"open_count": 0, "total_seconds": 0.0, "last_opened_at": "", "last_closed_at": ""}
        try:
            with open(self.summary_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {"open_count": 0, "total_seconds": 0.0, "last_opened_at": "", "last_closed_at": ""}

    def _save_summary(self, data):
        tmp = self.summary_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, self.summary_path)

    def _increment_open_count(self):
        data = self._load_summary()
        data["open_count"] = int(data.get("open_count", 0)) + 1
        data["last_opened_at"] = time.strftime("%d/%m/%Y %H:%M:%S")
        self._save_summary(data)
        return data["open_count"]

    def log(self, action: str, details: str = "", extra: Optional[dict] = None):
        record = {
            "data_hora": time.strftime("%d/%m/%Y %H:%M:%S"),
            "timestamp": time.time(),
            "sessao": self.session_id,
            "acao": action,
            "detalhes": details,
            "ip": self.ip,
            "maquina": self.host,
            "usuario": self.user,
            "sistema": platform.platform(),
            "diretorio": os.getcwd(),
            "abertura_numero": self.open_count,
            "tempo_sessao_segundos": round(time.time() - self.started_at, 2),
            "extra": extra or {},
        }
        os.makedirs(self.base_dir, exist_ok=True)
        with open(self.path, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    def close(self):
        duration = round(time.time() - self.started_at, 2)
        data = self._load_summary()
        data["total_seconds"] = round(float(data.get("total_seconds", 0.0)) + duration, 2)
        data["last_closed_at"] = time.strftime("%d/%m/%Y %H:%M:%S")
        self._save_summary(data)
        self.log("FECHAMENTO", f"Sistema encerrado. Tempo de uso: {duration}s", {"duracao_segundos": duration})

    def read_events(self, limit: int = 1000):
        if not os.path.exists(self.path):
            return []
        rows = []
        with open(self.path, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    rows.append(json.loads(line))
                except Exception:
                    pass
        return rows[-limit:]

    def summary_text(self):
        data = self._load_summary()
        total = float(data.get("total_seconds", 0.0))
        horas = int(total // 3600)
        minutos = int((total % 3600) // 60)
        segundos = int(total % 60)
        return (
            f"Aberturas: {data.get('open_count', 0)} | "
            f"Tempo total: {horas:02d}:{minutos:02d}:{segundos:02d} | "
            f"IP atual: {self.ip} | Máquina: {self.host} | Usuário: {self.user}"
        )

class SessionStore:
    def __init__(self, path: str):
        self.path = path

    def save(self, data: dict):
        tmp = self.path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, self.path)

    def load(self):
        if not os.path.exists(self.path):
            return None
        with open(self.path, "r", encoding="utf-8") as f:
            return json.load(f)

class WorkbookAdapter:
    def detect(self, path: str):
        wb = load_workbook(path)
        best = None
        for ws in wb.worksheets:
            for row_idx in range(1, min(ws.max_row, 15) + 1):
                headers = [normalize_text(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
                score, mapping = 0, {}
                for logical, options in HEADER_SYNONYMS.items():
                    exact = next((i for i, h in enumerate(headers, start=1) if h in options), None)
                    if exact:
                        mapping[logical] = exact
                        score += 4
                        continue
                    partial = next((i for i, h in enumerate(headers, start=1) if any(opt in h for opt in options)), None)
                    if partial:
                        mapping[logical] = partial
                        score += 1
                if "codigo" in mapping and "nome" in mapping:
                    candidate = (score, ws.title, row_idx, mapping)
                    if best is None or score > best[0]:
                        best = candidate

        if best is None:
            raise RuntimeError("Não consegui identificar automaticamente o cabeçalho da planilha.")

        _, sheet_name, header_row, mapping = best
        ws = wb[sheet_name]
        items = []

        for row in range(header_row + 1, ws.max_row + 1):
            codigo = normalize_code(ws.cell(row, mapping["codigo"]).value)
            nome = str(ws.cell(row, mapping["nome"]).value or "").strip()
            if not codigo and not nome:
                continue

            quantidade = to_float(ws.cell(row, mapping.get("quantidade", 0)).value) if mapping.get("quantidade") else 0.0
            disponivel = to_float(ws.cell(row, mapping.get("disponivel", 0)).value) if mapping.get("disponivel") else quantidade

            items.append(InventoryItem(
                codigo=codigo,
                nome=nome,
                marca=str(ws.cell(row, mapping.get("marca", 0)).value or "").strip() if mapping.get("marca") else "",
                quantidade=quantidade,
                disponivel=disponivel,
                unidade=str(ws.cell(row, mapping.get("unidade", 0)).value or "").strip() if mapping.get("unidade") else "",
                contagem_inventario=to_float(ws.cell(row, mapping.get("contagem", 0)).value) if mapping.get("contagem") else 0.0,
                linha_excel=row,
            ))

        headers_original = [str(ws.cell(header_row, c).value or "") for c in range(1, ws.max_column + 1)]
        preview = [headers_original]
        for row in range(header_row + 1, min(ws.max_row, header_row + 8) + 1):
            preview.append([str(ws.cell(row, c).value or "") for c in range(1, ws.max_column + 1)])

        meta = {
            "file_path": path,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "columns": mapping,
            "headers_original": headers_original,
            "items_count": len(items)
        }
        return items, meta, preview

    def create_blank_session(self):
        headers = ["Cod Produto", "Nome do Produto", "Marca", "Quantidade", "Disponível", "Unidade", "Contagem Inventário", "Diferença", "Última Leitura", "Total Scans", "Status", "Origem"]
        return {"file_path": "", "sheet_name": "Inventario", "header_row": 1, "columns": {"codigo": 1, "nome": 2, "marca": 3, "quantidade": 4, "disponivel": 5, "unidade": 6, "contagem": 7, "diferenca": 8}, "headers_original": headers, "items_count": 0, "mode": "blank", "created_at": time.strftime("%d/%m/%Y %H:%M:%S")}

    def _prepare_new_workbook(self, session: dict):
        wb = Workbook()
        ws = wb.active
        ws.title = session.get("sheet_name") or "Inventario"
        headers = ["Cod Produto", "Nome do Produto", "Marca", "Quantidade", "Disponível", "Unidade", "Contagem Inventário", "Diferença", "Última Leitura", "Total Scans", "Status", "Origem"]
        for col, head in enumerate(headers, start=1):
            ws.cell(1, col).value = head
        session["header_row"] = 1
        session["sheet_name"] = ws.title
        session["columns"] = {"codigo": 1, "nome": 2, "marca": 3, "quantidade": 4, "disponivel": 5, "unidade": 6, "contagem": 7, "diferenca": 8}
        return wb, ws

    def export_csv(self, session: dict, destination: str):
        items = [InventoryItem(**raw) for raw in session["items"]]
        for item in items:
            item.sincronizar_contagem_sem_planilha()
        headers = ["Cod Produto", "Nome do Produto", "Marca", "Quantidade", "Disponível", "Unidade", "Contagem Inventário", "Diferença", "Última Leitura", "Total Scans", "Status", "Origem"]
        with open(destination, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(headers)
            for item in sorted(items, key=lambda i: i.codigo):
                writer.writerow([item.codigo, item.nome, item.marca, item.quantidade, item.disponivel, item.unidade, item.contagem_inventario, item.diferenca, item.ultima_leitura, item.total_scans, item.status, item.origem])

    def _write_item_to_row(self, ws, row: int, columns: dict, item: InventoryItem, cont_col: int, dif_col: int):
        if columns.get("codigo"): ws.cell(row, columns["codigo"]).value = item.codigo
        if columns.get("nome"): ws.cell(row, columns["nome"]).value = item.nome
        if columns.get("marca"): ws.cell(row, columns["marca"]).value = item.marca
        if columns.get("quantidade"): ws.cell(row, columns["quantidade"]).value = item.quantidade
        if columns.get("disponivel"): ws.cell(row, columns["disponivel"]).value = item.disponivel
        if columns.get("unidade"): ws.cell(row, columns["unidade"]).value = item.unidade
        ws.cell(row, cont_col).value = item.contagem_inventario
        ws.cell(row, dif_col).value = item.diferenca
        if ws.max_column >= 9: ws.cell(row, 9).value = item.ultima_leitura
        if ws.max_column >= 10: ws.cell(row, 10).value = item.total_scans
        if ws.max_column >= 11: ws.cell(row, 11).value = item.status
        if ws.max_column >= 12: ws.cell(row, 12).value = item.origem

    def export(self, session: dict, destination: str):
        if session.get("mode") == "blank" or not session.get("file_path"):
            wb, ws = self._prepare_new_workbook(session)
        else:
            wb = load_workbook(session["file_path"])
            ws = wb[session["sheet_name"]]
        header_row, columns = session["header_row"], session["columns"]
        cont_col = columns.get("contagem")
        dif_col = columns.get("diferenca")
        last_col = ws.max_column
        if not cont_col:
            cont_col = last_col + 1
            columns["contagem"] = cont_col
            last_col = cont_col
        if not dif_col:
            dif_col = last_col + 1
            columns["diferenca"] = dif_col
        ws.cell(header_row, cont_col).value = "Contagem Inventário"
        ws.cell(header_row, dif_col).value = "Diferença"
        items = [InventoryItem(**raw) for raw in session["items"]]
        for item in items:
            item.sincronizar_contagem_sem_planilha()
        max_row = ws.max_row
        for item in items:
            if item.linha_excel and not item.novo_item:
                self._write_item_to_row(ws, item.linha_excel, columns, item, cont_col, dif_col)
            else:
                max_row += 1
                if max_row > header_row + 1:
                    self._clone_row_style(ws, max(max_row - 1, header_row + 1), max_row)
                self._write_item_to_row(ws, max_row, columns, item, cont_col, dif_col)
        for col in range(1, ws.max_column + 1):
            letter = ws.cell(header_row, col).column_letter
            title = str(ws.cell(header_row, col).value or "")
            ws.column_dimensions[letter].width = max(12, min(38, len(title) + 4))
        wb.save(destination)

    @staticmethod
    def _clone_row_style(ws, source_row: int, target_row: int):
        for col in range(1, ws.max_column + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)
            if src.has_style:
                dst._style = copy.copy(src._style)
                dst.font = copy.copy(src.font)
                dst.fill = copy.copy(src.fill)
                dst.border = copy.copy(src.border)
                dst.alignment = copy.copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy.copy(src.protection)

class AddProductDialog(tk.Toplevel):
    def __init__(self, master, codigo: str, contagem_sem_planilha: bool = False):
        super().__init__(master)
        self.title("Cadastrar novo item")
        self.contagem_sem_planilha = contagem_sem_planilha
        self.configure(bg=COLORS["panel"]); self.resizable(False, False)
        self.transient(master); self.grab_set(); self.result = None
        frm = tk.Frame(self, bg=COLORS["panel"], padx=16, pady=16); frm.pack(fill="both", expand=True)
        self.vars = {"codigo": tk.StringVar(value=codigo), "nome": tk.StringVar(), "marca": tk.StringVar(), "unidade": tk.StringVar(value="UN")}
        rows = [("Código","codigo"),("Nome","nome"),("Marca","marca"),("Unidade","unidade")]
        for i,(lab,key) in enumerate(rows):
            tk.Label(frm, text=lab, bg=COLORS["panel"], fg=COLORS["text"]).grid(row=i,column=0,sticky="w", pady=4)
            ent = tk.Entry(frm, textvariable=self.vars[key], width=34); ent.grid(row=i,column=1,sticky="ew",padx=(10,0), pady=4)
            if key == "codigo": ent.configure(state="readonly")
        btns = tk.Frame(frm, bg=COLORS["panel"]); btns.grid(row=len(rows), column=0, columnspan=2, sticky="e", pady=(12,0))
        tile_button(btns, text="Cancelar", command=self.destroy).pack(side="right", padx=6)
        tile_button(btns, text="Salvar", command=self._save, primary=True).pack(side="right")
        frm.columnconfigure(1, weight=1)

    def _save(self):
        nome = self.vars["nome"].get().strip()
        if not nome:
            messagebox.showwarning("Cadastro", "Informe o nome do produto.", parent=self); return
        self.result = InventoryItem(
            codigo=self.vars["codigo"].get().strip(),
            nome=nome,
            marca=self.vars["marca"].get().strip(),
            unidade=self.vars["unidade"].get().strip(),
            quantidade=0.0,
            disponivel=0.0,
            contagem_inventario=1.0,
            novo_item=True,
            contagem_sem_planilha=self.contagem_sem_planilha,
            ultima_leitura=time.strftime("%d/%m/%Y %H:%M:%S"),
            total_scans=1
        )
        self.result.sincronizar_contagem_sem_planilha()
        self.destroy()

class ImportPreviewDialog(tk.Toplevel):
    def __init__(self, master, file_name: str, meta: dict, preview_rows: List[List[str]]):
        super().__init__(master)
        self.title("Prévia da importação")
        self.configure(bg=COLORS["panel"]); self.geometry("1120x620")
        self.transient(master); self.grab_set(); self.result = False
        wrap = tk.Frame(self, bg=COLORS["panel"], padx=14, pady=14); wrap.pack(fill="both", expand=True)
        tk.Label(wrap, text="Prévia da planilha", font=("Segoe UI", 18, "bold"), bg=COLORS["panel"], fg=COLORS["text"]).pack(anchor="w")
        tk.Label(wrap, text=f"Arquivo: {os.path.basename(file_name)} | Aba: {meta['sheet_name']} | Cabeçalho: linha {meta['header_row']} | Itens: {meta['items_count']}", bg=COLORS["panel"], fg=COLORS["muted"]).pack(anchor="w", pady=(4,10))
        tree_frame = tk.Frame(wrap, bg=COLORS["white"], bd=1, relief="solid"); tree_frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(tree_frame, show="headings"); tree.pack(side="left", fill="both", expand=True)
        ysb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview); ysb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=ysb.set)
        headers = preview_rows[0] if preview_rows else []
        cols = [f"c{i}" for i in range(len(headers))]; tree["columns"] = cols
        for i, head in enumerate(headers):
            tree.heading(cols[i], text=head or f"Coluna {i+1}"); tree.column(cols[i], width=160, anchor="w")
        for row in preview_rows[1:]:
            tree.insert("", "end", values=row)
        btns = tk.Frame(wrap, bg=COLORS["panel"]); btns.pack(fill="x", pady=(14,0))
        tile_button(btns, text="CONTINUAR", command=self._continue, primary=True).pack(side="left", fill="x", expand=True, padx=(0,8))
        tile_button(btns, text="CANCELAR OPERAÇÃO", command=self._cancel, danger=True).pack(side="left", fill="x", expand=True, padx=(8,0))

    def _continue(self):
        self.result = True
        self.destroy()

    def _cancel(self):
        self.result = False
        self.destroy()

class StartupSplash:
    MAX_MS = 7700

    def __init__(self, root, on_finish):
        self.root = root
        self.on_finish = on_finish
        self.finished = False
        self.start_time = time.time()
        self.win = tk.Toplevel(root)
        self.win.overrideredirect(True)
        self.win.configure(bg="#05070a")
        self._fullscreen()
        try:
            self.win.attributes("-topmost", True)
        except Exception:
            pass
        self._build()
        self._animate()
        self.win.after(self.MAX_MS, self.finish)

    def _fullscreen(self):
        self.win.update_idletasks()
        width = self.win.winfo_screenwidth()
        height = self.win.winfo_screenheight()
        self.win.geometry(f"{width}x{height}+0+0")
        try:
            self.win.attributes("-fullscreen", True)
        except Exception:
            pass

    def _build(self):
        self.screen_w = self.win.winfo_screenwidth()
        self.screen_h = self.win.winfo_screenheight()
        self.canvas = tk.Canvas(self.win, width=self.screen_w, height=self.screen_h, bg="#05070a", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        self.cx = self.screen_w // 2
        self.cy = self.screen_h // 2
        self.panel_w = min(760, max(520, int(self.screen_w * 0.52)))
        self.panel_h = min(430, max(340, int(self.screen_h * 0.42)))
        x1 = self.cx - self.panel_w // 2
        y1 = self.cy - self.panel_h // 2
        x2 = self.cx + self.panel_w // 2
        y2 = self.cy + self.panel_h // 2
        self.canvas.create_rectangle(x1, y1, x2, y2, outline="#1f2a3b", fill="#080d14", width=1)
        self.canvas.create_rectangle(x1, y1, x2, y1 + 58, outline="", fill="#0f3d91")
        self.canvas.create_text(x1 + 28, y1 + 29, text="TÖNCHTORY", fill="#ffffff", font=("Segoe UI Light", 24), anchor="w")
        self.canvas.create_text(x2 - 28, y1 + 30, text="POWERED BY GABRIEL M PERDIGAO", fill="#dbeafe", font=("Segoe UI", 10), anchor="e")
        self.tri_y_base = y1 + 168
        self.tri_y_current = self.tri_y_base
        self.tri_group = []
        self._draw_triangle(self.cx, self.tri_y_base)
        self.title_text = self.canvas.create_text(self.cx, y1 + 294, text="", fill="#e6eeff", font=("Segoe UI Semibold", 18), anchor="center")
        self.sub_text = self.canvas.create_text(self.cx, y1 + 328, text="", fill="#9fb2c8", font=("Segoe UI", 10), anchor="center")
        self.bar_bg = self.canvas.create_rectangle(x1 + 74, y2 - 64, x2 - 74, y2 - 54, outline="", fill="#192334")
        self.bar = self.canvas.create_rectangle(x1 + 74, y2 - 64, x1 + 75, y2 - 54, outline="", fill="#2d89ef")
        self.canvas.create_text(x1 + 24, y2 - 24, text="TÖNCH Technology Solutions", fill="#70859f", font=("Segoe UI", 9), anchor="w")

    def _draw_triangle(self, cx, cy):
        c = self.canvas
        self.tri_group = [
            c.create_polygon(cx, cy - 88, cx + 90, cy + 72, cx - 90, cy + 72, outline="#dbeafe", fill="", width=3),
            c.create_polygon(cx, cy - 58, cx + 62, cy + 56, cx - 62, cy + 56, outline="#2d89ef", fill="", width=2),
            c.create_oval(cx - 40, cy - 18, cx + 40, cy + 24, outline="#ffffff", width=3),
            c.create_oval(cx - 16, cy - 13, cx + 16, cy + 19, fill="#2d89ef", outline="#dbeafe", width=2),
            c.create_oval(cx - 5, cy - 2, cx + 5, cy + 8, fill="#05070a", outline=""),
            c.create_line(cx, cy - 112, cx, cy - 92, fill="#2d89ef", width=2),
            c.create_line(cx, cy + 76, cx, cy + 106, fill="#2d89ef", width=2),
            c.create_line(cx - 122, cy, cx - 92, cy, fill="#2d89ef", width=2),
            c.create_line(cx + 92, cy, cx + 122, cy, fill="#2d89ef", width=2),
            c.create_line(cx - 84, cy - 84, cx - 64, cy - 64, fill="#2d89ef", width=2),
            c.create_line(cx + 84, cy - 84, cx + 64, cy - 64, fill="#2d89ef", width=2),
            c.create_line(cx - 84, cy + 84, cx - 64, cy + 64, fill="#2d89ef", width=2),
            c.create_line(cx + 84, cy + 84, cx + 64, cy + 64, fill="#2d89ef", width=2),
        ]

    def _move_triangle_to(self, y):
        delta = y - self.tri_y_current
        self.tri_y_current = y
        if delta:
            for item in self.tri_group:
                self.canvas.move(item, 0, delta)

    def _animate(self):
        if self.finished:
            return
        import math
        elapsed = time.time() - self.start_time
        self._move_triangle_to(self.tri_y_base + int(10 * math.sin(elapsed * 4.0)))
        msg = "TÖNCHTORY - Iniciando o sistema"
        sub = "Preparando leitura, cadastro, contagem e exportação..."
        self.canvas.itemconfigure(self.title_text, text=msg[:min(len(msg), int(elapsed * 18))])
        if elapsed > 1.1:
            self.canvas.itemconfigure(self.sub_text, text=sub[:min(len(sub), int((elapsed - 1.1) * 24))])
        try:
            self.win.attributes("-alpha", 0.78 + 0.22 * abs(math.sin(elapsed * 2.7)))
        except Exception:
            pass
        progress = min(1.0, elapsed / (self.MAX_MS / 1000.0))
        x1, y1, x2, y2 = self.canvas.coords(self.bar_bg)
        self.canvas.coords(self.bar, x1, y1, x1 + max(1, int((x2 - x1) * progress)), y2)
        self.win.after(45, self._animate)

    def finish(self):
        if self.finished:
            return
        self.finished = True
        try:
            self.win.destroy()
        except Exception:
            pass
        self.on_finish()

class App:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self._apply_fullscreen(); self.root.minsize(1024, 680)
        self.root.configure(bg=COLORS["bg"])
        self.adapter = WorkbookAdapter()
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.store = SessionStore(os.path.join(self.base_dir, SESSION_FILE))
        self.audit = AuditManager(self.base_dir)
        self.audit.log("ABERTURA", "Sistema aberto")
        self.session = None
        self.items: List[InventoryItem] = []
        self.undo_stack = []
        self.selected_code = None
        self.audit_win = None
        self._build_ui()
        self._load_saved_session()
        self._ensure_focus()

    def _apply_fullscreen(self):
        try:
            self.root.attributes("-fullscreen", True)
        except Exception:
            pass
        try:
            self.root.state("zoomed")
        except Exception:
            pass
        self.root.update_idletasks()

    def force_exit(self):
        try:
            self._persist()
            self.audit.close()
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass
        os._exit(0)

    def _build_ui(self):
        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        style.configure("Treeview", background=COLORS["white"], fieldbackground=COLORS["white"], foreground=COLORS["text"], rowheight=28, bordercolor=COLORS["border"], borderwidth=0, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", background=COLORS["header"], foreground=COLORS["text"], relief="flat", font=("Segoe UI", 9, "bold"))
        style.configure("Vertical.TScrollbar", background=COLORS["panel"], troughcolor=COLORS["bg"], bordercolor=COLORS["border"], arrowcolor=COLORS["text"])
        style.configure("Horizontal.TScrollbar", background=COLORS["panel"], troughcolor=COLORS["bg"], bordercolor=COLORS["border"], arrowcolor=COLORS["text"])
        style.map("Treeview", background=[("selected", COLORS["select_bg"])], foreground=[("selected", COLORS["select_fg"])])

        title = tk.Frame(self.root, bg=COLORS["header"], bd=1, relief="solid"); title.pack(fill="x", padx=8, pady=8)
        left = tk.Frame(title, bg=COLORS["header"]); left.pack(side="left", fill="x", expand=True, padx=14, pady=12)
        tk.Label(left, text=APP_TITLE, font=("Segoe UI", 17, "bold"), bg=COLORS["header"], fg=COLORS["text"]).pack(anchor="w")
        tk.Label(left, text="Inventário, contagem, cadastro e exportação em uma única operação.", bg=COLORS["header"], fg=COLORS["muted"]).pack(anchor="w", pady=(4,0))
        self.lbl_scanner = tk.Label(title, text="Leitor: aguardando atividade", bg="#eaf4ff", fg=COLORS["blue_dark"], bd=1, relief="solid", padx=16, pady=10); self.lbl_scanner.pack(side="right", padx=14, pady=12)

        toolbar = tk.Frame(self.root, bg=COLORS["panel"], bd=1, relief="solid"); toolbar.pack(fill="x", padx=8)
        for txt, cmd in [("Importar Planilha", self.import_workbook), ("Nova Contagem sem Planilha", self.new_blank_count), ("Continuar Sessão", self.continue_session), ("Desfazer Última", self.undo_last), ("Recontar Selecionado", self.recount_selected), ("Zerar Selecionado", self.zero_selected)]:
            tile_button(toolbar, text=txt, command=cmd).pack(side="left", padx=6, pady=6)
        tile_button(toolbar, text="SAIR", command=self.force_exit, danger=True).pack(side="right", padx=6, pady=6)
        tile_button(toolbar, text="Auditoria", command=self.open_audit).pack(side="right", padx=6, pady=6)
        tile_button(toolbar, text="Salvar Inventário", command=self.save_inventory, primary=True).pack(side="right", padx=6, pady=6)

        summary = tk.Frame(self.root, bg=COLORS["white"], bd=1, relief="solid"); summary.pack(fill="x", padx=8, pady=(8,4))
        self.var_file = tk.StringVar(value="Arquivo: nenhum"); self.var_sheet = tk.StringVar(value="Aba: -"); self.var_rows = tk.StringVar(value="Linhas carregadas: 0")
        self.var_counted = tk.StringVar(value="Itens com contagem: 0"); self.var_diff = tk.StringVar(value="Divergências: 0"); self.var_map = tk.StringVar(value="Mapeamento: -"); self.var_last = tk.StringVar(value="Último scan: -")
        row1 = tk.Frame(summary, bg=COLORS["white"]); row1.pack(fill="x", padx=12, pady=(8,3))
        row2 = tk.Frame(summary, bg=COLORS["white"]); row2.pack(fill="x", padx=12, pady=(0,8))
        for var in [self.var_file, self.var_sheet, self.var_rows, self.var_counted, self.var_diff]:
            tk.Label(row1, textvariable=var, bg=COLORS["white"], fg=COLORS["text"]).pack(side="left", padx=(0,18))
        for var in [self.var_map, self.var_last]:
            tk.Label(row2, textvariable=var, bg=COLORS["white"], fg=COLORS["muted"]).pack(side="left", padx=(0,18))

        scan_box = tk.Frame(self.root, bg=COLORS["panel"], bd=1, relief="solid"); scan_box.pack(fill="x", padx=8, pady=(4,6))
        tk.Label(scan_box, text="Entrada de scanner / teclado", bg=COLORS["panel"], fg=COLORS["text"], font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(10,4))
        self.scan_var = tk.StringVar()
        self.scan_entry = tk.Entry(scan_box, textvariable=self.scan_var, font=("Consolas", 22), relief="flat", bd=0, bg="#0d1118", fg="#ffffff", insertbackground="#ffffff")
        self.scan_entry.pack(fill="x", padx=12, pady=(0,10))
        self.scan_entry.bind("<Return>", self.on_scan_enter)
        self.scan_entry.bind("<KeyRelease>", lambda e: self.lbl_scanner.config(text="Leitor: atividade detectada"))

        grid_frame = tk.Frame(self.root, bg=COLORS["white"], bd=1, relief="solid"); grid_frame.pack(fill="both", expand=True, padx=8, pady=(0,8))
        cols = ("codigo","nome","marca","quantidade","disponivel","unidade","contagem","diferenca","ultima","scans","status","origem")
        self.tree = ttk.Treeview(grid_frame, columns=cols, show="headings"); self.tree.pack(side="left", fill="both", expand=True)
        ysb = ttk.Scrollbar(grid_frame, orient="vertical", command=self.tree.yview); ysb.pack(side="right", fill="y")
        xsb = ttk.Scrollbar(self.root, orient="horizontal", command=self.tree.xview); xsb.pack(fill="x", padx=8)
        self.tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        heads = {"codigo":"Cod Produto","nome":"Nome do Produto","marca":"Marca","quantidade":"Quantidade","disponivel":"Disponível","unidade":"Unidade","contagem":"Contagem Inventário","diferenca":"Diferença","ultima":"Última Leitura","scans":"Total Scans","status":"Status","origem":"Origem"}
        widths = {"codigo":110,"nome":380,"marca":140,"quantidade":95,"disponivel":95,"unidade":80,"contagem":145,"diferenca":95,"ultima":145,"scans":95,"status":95,"origem":80}
        for c in cols:
            self.tree.heading(c, text=heads[c]); self.tree.column(c, width=widths[c], anchor="w")
        self.tree.tag_configure("odd", background=COLORS["alt"]); self.tree.tag_configure("positive", foreground=COLORS["blue_dark"]); self.tree.tag_configure("negative", foreground="#ff6b6b"); self.tree.tag_configure("ok", foreground="#7ee787")
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        bottom = tk.Frame(self.root, bg=COLORS["panel"], bd=1, relief="solid"); bottom.pack(fill="x", padx=8, pady=(0,8))
        self.info = tk.Label(bottom, text="Pronto. Importe uma planilha ou inicie uma nova contagem sem planilha importada.", bg=COLORS["panel"], fg=COLORS["muted"], anchor="w", padx=12, pady=10); self.info.pack(fill="x")

    def open_audit(self):
        """Abre a auditoria em janela própria e mantém ela acima do app fullscreen."""
        try:
            if getattr(self, "audit_win", None) is not None and self.audit_win.winfo_exists():
                self.audit_win.deiconify()
                self.audit_win.lift()
                self.audit_win.focus_force()
                return
        except Exception:
            self.audit_win = None

        try:
            self.audit.log("AUDITORIA", "Menu de auditoria aberto")
        except Exception:
            pass

        win = tk.Toplevel(self.root)
        self.audit_win = win
        win.title("Auditoria do Sistema")
        win.configure(bg=COLORS["panel"])
        win.minsize(980, 620)

        def close_audit():
            try:
                self.audit_win = None
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass
            try:
                self.root.after(300, self._ensure_focus)
            except Exception:
                pass

        win.protocol("WM_DELETE_WINDOW", close_audit)

        try:
            win.transient(self.root)
        except Exception:
            pass
        try:
            win.attributes("-topmost", True)
        except Exception:
            pass
        try:
            win.state("zoomed")
        except Exception:
            center_window(win, 1180, 720)
        try:
            win.lift()
            win.focus_force()
            win.after(700, lambda: win.attributes("-topmost", False) if win.winfo_exists() else None)
        except Exception:
            pass

        try:
            top = tk.Frame(win, bg=COLORS["panel"])
            top.pack(fill="x", padx=14, pady=(12, 8))
            tk.Label(top, text="Auditoria Enterprise", bg=COLORS["panel"], fg=COLORS["text"], font=("Segoe UI", 18, "bold")).pack(anchor="w")
            tk.Label(top, text=str(self.audit.summary_text()), bg=COLORS["panel"], fg=COLORS["muted"], font=("Segoe UI", 10)).pack(anchor="w", pady=(4, 0))

            grid = tk.Frame(win, bg=COLORS["white"], bd=1, relief="solid")
            grid.pack(fill="both", expand=True, padx=14, pady=(0, 10))

            cols = ("data_hora", "acao", "detalhes", "ip", "maquina", "usuario", "tempo", "sessao")
            tree = ttk.Treeview(grid, columns=cols, show="headings")
            tree.pack(side="left", fill="both", expand=True)

            ysb = ttk.Scrollbar(grid, orient="vertical", command=tree.yview)
            ysb.pack(side="right", fill="y")
            tree.configure(yscrollcommand=ysb.set)

            heads = {"data_hora": "Data/Hora", "acao": "Ação", "detalhes": "Detalhes", "ip": "IP", "maquina": "Máquina", "usuario": "Usuário", "tempo": "Tempo sessão", "sessao": "Sessão"}
            widths = {"data_hora": 145, "acao": 150, "detalhes": 440, "ip": 120, "maquina": 150, "usuario": 120, "tempo": 105, "sessao": 110}
            for c in cols:
                tree.heading(c, text=heads[c])
                tree.column(c, width=widths[c], anchor="w")

            try:
                events = self.audit.read_events(1000)
            except Exception:
                events = []

            for ev in reversed(events):
                if not isinstance(ev, dict):
                    continue
                tree.insert("", "end", values=(
                    str(ev.get("data_hora", "")),
                    str(ev.get("acao", "")),
                    str(ev.get("detalhes", "")),
                    str(ev.get("ip", "")),
                    str(ev.get("maquina", "")),
                    str(ev.get("usuario", "")),
                    str(ev.get("tempo_sessao_segundos", "")),
                    str(ev.get("sessao", "")),
                ))

            bottom = tk.Frame(win, bg=COLORS["panel"])
            bottom.pack(fill="x", padx=14, pady=(0, 14))

            def export_audit():
                out = filedialog.asksaveasfilename(parent=win, defaultextension=".csv", initialfile="auditoria_tonchtory.csv", filetypes=[("CSV", "*.csv")])
                if not out:
                    return
                try:
                    events_export = self.audit.read_events(100000)
                    with open(out, "w", newline="", encoding="utf-8-sig") as f:
                        writer = csv.writer(f, delimiter=";")
                        writer.writerow(["Data/Hora", "Sessão", "Ação", "Detalhes", "IP", "Máquina", "Usuário", "Sistema", "Diretório", "Abertura", "Tempo sessão", "Extra"])
                        for ev in events_export:
                            if not isinstance(ev, dict):
                                continue
                            writer.writerow([ev.get("data_hora", ""), ev.get("sessao", ""), ev.get("acao", ""), ev.get("detalhes", ""), ev.get("ip", ""), ev.get("maquina", ""), ev.get("usuario", ""), ev.get("sistema", ""), ev.get("diretorio", ""), ev.get("abertura_numero", ""), ev.get("tempo_sessao_segundos", ""), json.dumps(ev.get("extra", {}), ensure_ascii=False)])
                    try:
                        self.audit.log("AUDITORIA_EXPORTADA", out)
                    except Exception:
                        pass
                    messagebox.showinfo("Auditoria", f"Auditoria exportada com sucesso.\n\n{out}", parent=win)
                except Exception as e:
                    messagebox.showerror("Auditoria", f"Não foi possível exportar a auditoria.\n\n{e}", parent=win)

            tile_button(bottom, text="Exportar Auditoria CSV", command=export_audit, primary=True).pack(side="left")
            tile_button(bottom, text="Fechar", command=close_audit).pack(side="right")

        except Exception as e:
            for child in win.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            box = tk.Frame(win, bg=COLORS["panel"])
            box.pack(fill="both", expand=True, padx=24, pady=24)
            tk.Label(box, text="Auditoria não conseguiu montar a grade, mas o sistema continua aberto.", bg=COLORS["panel"], fg=COLORS["text"], font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 10))
            tk.Label(box, text=str(e), bg=COLORS["panel"], fg=COLORS["muted"], wraplength=900, justify="left").pack(anchor="w")
            tile_button(box, text="Fechar", command=close_audit).pack(anchor="w", pady=18)

    def _ensure_focus(self):
        try:
            if getattr(self, "audit_win", None) is not None and self.audit_win.winfo_exists():
                return
        except Exception:
            self.audit_win = None
        try:
            self.scan_entry.focus_set()
        except Exception:
            pass
        self.root.after(900, self._ensure_focus)

    def _load_saved_session(self):
        data = self.store.load()
        if data:
            self.session = data
            self.items = [InventoryItem(**raw) for raw in data["items"]]
            for item in self.items:
                item.sincronizar_contagem_sem_planilha()
            self.undo_stack = data.get("undo_stack", [])
            self.selected_code = self.items[0].codigo if self.items else None
            self._refresh()

    def _set_info(self, text):
        self.info.config(text=text)

    def _session_dict(self):
        return {**self.session, "items":[asdict(i) for i in self.items], "undo_stack":self.undo_stack[-500:]}

    def _persist(self):
        if self.session:
            self.store.save(self._session_dict())

    def import_workbook(self):
        path = filedialog.askopenfilename(title="Selecione a planilha", filetypes=[("Excel", "*.xlsx *.xlsm *.xltx *.xltm")])
        if not path: return
        try:
            items, meta, preview = self.adapter.detect(path)
        except Exception as e:
            messagebox.showerror("Importação", f"Falha ao importar a planilha.\n\n{e}")
            return
        dlg = ImportPreviewDialog(self.root, path, meta, preview)
        self.root.wait_window(dlg)
        if not dlg.result:
            self.session = None; self.items = []; self.undo_stack = []; self.selected_code = None
            self._refresh(clear_only=True)
            self._set_info("Importação cancelada. Prévia fechada e planilha removida.")
            return
        self.session = meta; self.items = items; self.undo_stack = []; self.selected_code = self.items[0].codigo if self.items else None
        self.audit.log("IMPORTACAO_PLANILHA", os.path.basename(path), {"arquivo": path, "itens": len(items), "aba": meta.get("sheet_name")})
        self._persist(); self._refresh(); self._set_info("Planilha importada e pronta para o trabalho.")

    def new_blank_count(self):
        if self.items and not messagebox.askyesno("Nova contagem", "Iniciar uma nova contagem em branco? A sessão atual será substituída na memória local."):
            return
        self.session = self.adapter.create_blank_session()
        self.items = []
        self.undo_stack = []
        self.selected_code = None
        self.audit.log("NOVA_CONTAGEM_SEM_PLANILHA", "Sessão em branco iniciada")
        self._persist()
        self._refresh()
        self._set_info("Nova contagem sem planilha iniciada. Bipe ou digite um código para cadastrar e contar.")

    def continue_session(self):
        data = self.store.load()
        if not data:
            messagebox.showinfo("Sessão", "Nenhuma sessão salva encontrada."); return
        self.session = data; self.items = [InventoryItem(**raw) for raw in data["items"]]
        for item in self.items:
            item.sincronizar_contagem_sem_planilha()
        self.undo_stack = data.get("undo_stack", []); self.selected_code = self.items[0].codigo if self.items else None
        self.audit.log("CONTINUAR_SESSAO", "Sessão local restaurada", {"itens": len(self.items)})
        self._refresh(); self._set_info("Sessão continuada com sucesso.")

    def on_scan_enter(self, event=None):
        code = normalize_code(self.scan_var.get()); self.scan_var.set("")
        if not code: return
        if not self.session:
            if messagebox.askyesno("Leitura", "Nenhuma planilha carregada. Deseja iniciar uma NOVA CONTAGEM SEM PLANILHA IMPORTADA?"):
                self.new_blank_count()
            else:
                return
        item = next((i for i in self.items if i.codigo == code), None)
        if item is None:
            if not messagebox.askyesno("Não cadastrado", f"CÓDIGO {code} NÃO CADASTRADO.\n\nDeseja cadastrá-lo?"): return
            is_blank_mode = bool(self.session and (self.session.get("mode") == "blank" or not self.session.get("file_path")))
            dlg = AddProductDialog(self.root, code, contagem_sem_planilha=is_blank_mode); self.root.wait_window(dlg)
            if dlg.result is None: return
            item = dlg.result; self.items.append(item)
            self.audit.log("CADASTRO_NOVO_ITEM", f"{item.codigo} - {item.nome}", {"planilha_importada": not bool(self.session and (self.session.get("mode") == "blank" or not self.session.get("file_path")))})
            self._set_info(f"Novo produto {item.codigo} cadastrado e contado automaticamente."); self.lbl_scanner.config(text="Leitor: novo item cadastrado via scanner")
        else:
            item.contagem_inventario += 1
            item.sincronizar_contagem_sem_planilha()
            item.total_scans += 1
            item.ultima_leitura = time.strftime("%d/%m/%Y %H:%M:%S")
            self._set_info(f"Leitura registrada: {item.codigo} | {item.nome}")
            self.lbl_scanner.config(text="Leitor: scanner em operação")
        self.audit.log("SCAN", f"{item.codigo} - {item.nome}", {"contagem": item.contagem_inventario, "quantidade": item.quantidade, "disponivel": item.disponivel, "novo_item": item.novo_item})
        self.undo_stack.append((item.codigo, 1)); self.selected_code = item.codigo; self.var_last.set(f"Último scan: {item.codigo} | {item.nome}")
        self._persist(); self._refresh(select_code=item.codigo)

    def undo_last(self):
        if not self.undo_stack:
            messagebox.showinfo("Desfazer", "Nada para desfazer."); return
        codigo, delta = self.undo_stack.pop()
        item = next((i for i in self.items if i.codigo == codigo), None)
        if not item: return
        item.contagem_inventario = max(0, item.contagem_inventario - delta)
        item.sincronizar_contagem_sem_planilha()
        item.total_scans = max(0, item.total_scans - delta)
        item.ultima_leitura = time.strftime("%d/%m/%Y %H:%M:%S")
        self.audit.log("DESFAZER_SCAN", codigo, {"contagem": item.contagem_inventario})
        self.selected_code = item.codigo; self._persist(); self._refresh(select_code=item.codigo); self._set_info(f"Última leitura desfeita para {codigo}.")

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        if selected: self.selected_code = selected[0]

    def _get_selected_item(self):
        if not self.selected_code:
            messagebox.showinfo("Seleção", "Selecione um item na grade."); return None
        return next((i for i in self.items if i.codigo == self.selected_code), None)

    def recount_selected(self):
        item = self._get_selected_item()
        if not item: return
        popup = tk.Toplevel(self.root); popup.title("Recontar item"); popup.configure(bg=COLORS["panel"]); popup.transient(self.root); popup.grab_set()
        tk.Label(popup, text=f"{item.codigo} - {item.nome}", bg=COLORS["panel"], fg=COLORS["text"]).pack(anchor="w", padx=16, pady=(16,6))
        var = tk.StringVar(value=fmt_number(item.contagem_inventario)); ent = tk.Entry(popup, textvariable=var, width=18); ent.pack(anchor="w", padx=16); ent.focus_set()
        def apply():
            item.contagem_inventario = max(0, to_float(var.get()))
            item.sincronizar_contagem_sem_planilha()
            item.ultima_leitura = time.strftime("%d/%m/%Y %H:%M:%S")
            self.audit.log("RECONTAR_ITEM", f"{item.codigo} - {item.nome}", {"contagem": item.contagem_inventario, "quantidade": item.quantidade, "disponivel": item.disponivel})
            popup.destroy(); self._persist(); self._refresh(select_code=item.codigo); self._set_info(f"Contagem reajustada para {item.codigo}.")
        tile_button(popup, text="Aplicar", command=apply, primary=True).pack(anchor="e", padx=16, pady=16)

    def zero_selected(self):
        item = self._get_selected_item()
        if not item: return
        if not messagebox.askyesno("Zerar", f"Zerar a contagem de {item.codigo}?"): return
        item.contagem_inventario = 0
        item.sincronizar_contagem_sem_planilha()
        item.ultima_leitura = time.strftime("%d/%m/%Y %H:%M:%S")
        self.audit.log("ZERAR_ITEM", f"{item.codigo} - {item.nome}")
        self._persist(); self._refresh(select_code=item.codigo); self._set_info(f"Contagem zerada para {item.codigo}.")

    def save_inventory(self):
        if not self.session:
            messagebox.showwarning("Salvar", "Nenhuma sessão carregada."); return
        base = os.path.splitext(os.path.basename(self.session.get("file_path") or "Nova Contagem"))[0]
        name = base + " - Inventario Final.xlsx"
        out = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=name, filetypes=[("Excel", "*.xlsx"), ("CSV separado por ponto e vírgula", "*.csv")])
        if not out: return
        self._persist()
        try:
            if out.lower().endswith(".csv"):
                self.adapter.export_csv(self._session_dict(), out)
            else:
                self.adapter.export(self._session_dict(), out)
            self._set_info(f"Inventário salvo em {out}"); messagebox.showinfo("Concluído", f"Inventário salvo com sucesso.\n\n{out}")
        except Exception as e:
            messagebox.showerror("Salvar", f"Falha ao salvar a planilha.\n\n{e}")

    def _refresh(self, select_code=None, clear_only=False):
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        if clear_only or not self.session:
            self.var_file.set("Arquivo: nenhum"); self.var_sheet.set("Aba: -"); self.var_rows.set("Linhas carregadas: 0"); self.var_counted.set("Itens com contagem: 0"); self.var_diff.set("Divergências: 0"); self.var_map.set("Mapeamento: -"); self.var_last.set("Último scan: -"); return
        for idx, item in enumerate(sorted(self.items, key=lambda i: i.codigo)):
            tags = ["odd"] if idx % 2 else []
            tags.append("positive" if item.diferenca > 0 else "negative" if item.diferenca < 0 else "ok")
            self.tree.insert("", "end", iid=item.codigo, tags=tuple(tags), values=(item.codigo, item.nome, item.marca, fmt_number(item.quantidade), fmt_number(item.disponivel), item.unidade, fmt_number(item.contagem_inventario), fmt_number(item.diferenca), item.ultima_leitura, item.total_scans, item.status, item.origem))
        code = select_code or self.selected_code
        if code and self.tree.exists(code):
            self.tree.selection_set(code); self.tree.focus(code); self.tree.see(code); self.selected_code = code
        if self.session.get("mode") == "blank" or not self.session.get("file_path"):
            self.var_file.set("Arquivo: NOVA CONTAGEM SEM PLANILHA IMPORTADA")
        else:
            self.var_file.set(f"Arquivo: {os.path.basename(self.session['file_path'])}")
        self.var_sheet.set(f"Aba: {self.session['sheet_name']}"); self.var_rows.set(f"Linhas carregadas: {len(self.items)}"); self.var_counted.set(f"Itens com contagem: {sum(1 for i in self.items if i.contagem_inventario > 0)}"); self.var_diff.set(f"Divergências: {sum(1 for i in self.items if i.diferenca != 0)}"); self.var_map.set("Mapeamento: " + ", ".join(f"{k}:{v}" for k, v in sorted(self.session["columns"].items())))

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def main():
    root = tk.Tk()
    root.withdraw()
    try:
        icon_path = resource_path("icon.ico")
        root.iconbitmap(icon_path)
    except Exception:
        pass

    def launch_app():
        App(root)
        root.deiconify()
        try:
            root.lift()
            root.focus_force()
            root.attributes("-fullscreen", True)
        except Exception:
            pass

    StartupSplash(root, launch_app)
    root.mainloop()

if __name__ == "__main__":
    main()
